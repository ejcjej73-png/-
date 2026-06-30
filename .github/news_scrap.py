# -*- coding: utf-8 -*-
"""
카드·외환 뉴스 스크랩 프로그램
================================
매일 실행하면 '카드' 및 '외환' 관련 신문기사를 검색해
날짜별 시트로 정리된 엑셀 파일(.xlsx)에 자동으로 모아 줍니다.

- 1순위: 네이버 뉴스 검색 API (CLIENT_ID/SECRET 입력 시, 한국 언론 커버리지 우수)
- 2순위: 구글 뉴스 RSS (키 없이도 즉시 동작 / 보완용)

사용법:  python news_scrap.py
일일 자동 실행은 README 참고(윈도우 작업 스케줄러 / cron).
"""

import os
import re
import sys
import html
import time
import datetime as dt
import configparser
from urllib.parse import quote

import requests
import feedparser
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 1. 설정 (이 부분만 수정하면 됩니다)
# ============================================================

# 네이버 검색 API 키 — https://developers.naver.com 에서 무료 발급
# 비워 두면 구글 뉴스 RSS만 사용합니다(키 없이도 동작).
NAVER_CLIENT_ID = ""
NAVER_CLIENT_SECRET = ""

# 검색 키워드 (분류별로 자유롭게 추가/삭제)
KEYWORDS = {
    "카드": [
        "신용카드", "체크카드", "카드사", "카드 수수료",
        "해외결제 카드", "간편결제", "카드 연체",
    ],
    "외환": [
        "외환", "환율", "원달러 환율", "해외송금",
        "외국환거래법", "외화예금", "환전",
    ],
}

# 며칠 전 기사까지 수집할지 (1 = 오늘 기사만, 2 = 어제+오늘 …)
DAYS_BACK = 1

# 키워드당 검색 건수(네이버 최대 100)
MAX_PER_KEYWORD = 30

# 결과 저장 폴더 및 파일명
OUTPUT_DIR = "."
OUTPUT_FILE = "카드외환_뉴스스크랩.xlsx"

# ============================================================
# 2. 내부 로직 (수정 불필요)
# ============================================================

KST = dt.timezone(dt.timedelta(hours=9))


def _base_dir():
    """실행 파일(.exe) 또는 스크립트가 있는 폴더."""
    if getattr(sys, "frozen", False):          # PyInstaller exe로 실행 중
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def load_external_config(filename="설정.ini"):
    """같은 폴더에 '설정.ini'가 있으면 그 값으로 덮어씁니다(없으면 기본값 사용).
    → EXE로 만든 뒤에도 파이썬 없이 키워드·설정을 바꿀 수 있습니다."""
    global NAVER_CLIENT_ID, NAVER_CLIENT_SECRET, DAYS_BACK
    global MAX_PER_KEYWORD, OUTPUT_FILE, KEYWORDS
    path = os.path.join(_base_dir(), filename)
    if not os.path.exists(path):
        return
    try:
        cp = configparser.ConfigParser()
        cp.optionxform = str                   # 키 대소문자/한글 보존
        cp.read(path, encoding="utf-8")
        s = cp["설정"] if cp.has_section("설정") else {}
        NAVER_CLIENT_ID = s.get("naver_id", NAVER_CLIENT_ID).strip()
        NAVER_CLIENT_SECRET = s.get("naver_secret", NAVER_CLIENT_SECRET).strip()
        DAYS_BACK = int(s.get("days_back", DAYS_BACK))
        MAX_PER_KEYWORD = int(s.get("max_per_keyword", MAX_PER_KEYWORD))
        OUTPUT_FILE = s.get("output_file", OUTPUT_FILE).strip()
        kw = {}
        if s.get("카드_키워드", "").strip():
            kw["카드"] = [w.strip() for w in s.get("카드_키워드").split(",") if w.strip()]
        if s.get("외환_키워드", "").strip():
            kw["외환"] = [w.strip() for w in s.get("외환_키워드").split(",") if w.strip()]
        if kw:
            KEYWORDS = kw
        print(f"· 설정 파일 적용: {path}")
    except Exception as e:
        print(f"· 설정 파일 읽기 오류(기본값 사용): {e}")


def _clean(text: str) -> str:
    """HTML 태그·엔티티 제거."""
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", "", text)
    text = html.unescape(text)
    return re.sub(r"\s+", " ", text).strip()


def _within_range(pub: dt.datetime, days_back: int) -> bool:
    if pub is None:
        return True  # 날짜를 못 읽으면 일단 포함
    today = dt.datetime.now(KST).date()
    return (today - pub.astimezone(KST).date()).days < days_back


def fetch_naver(keyword: str, category: str, display: int):
    """네이버 뉴스 검색 API. 키가 없으면 빈 리스트 반환."""
    if not (NAVER_CLIENT_ID and NAVER_CLIENT_SECRET):
        return []
    url = "https://openapi.naver.com/v1/search/news.json"
    headers = {
        "X-Naver-Client-Id": NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }
    params = {"query": keyword, "display": min(display, 100), "sort": "date"}
    out = []
    try:
        r = requests.get(url, headers=headers, params=params, timeout=10)
        r.raise_for_status()
        for it in r.json().get("items", []):
            pub = None
            try:
                pub = dt.datetime.strptime(it["pubDate"], "%a, %d %b %Y %H:%M:%S %z")
            except Exception:
                pass
            out.append({
                "category": category,
                "keyword": keyword,
                "title": _clean(it.get("title")),
                "press": "",  # 네이버 API는 언론사명을 직접 주지 않음(링크로 확인)
                "date": pub,
                "summary": _clean(it.get("description")),
                "link": it.get("originallink") or it.get("link", ""),
                "source": "네이버",
            })
    except Exception as e:
        print(f"  [네이버 오류] {keyword}: {e}")
    return out


def fetch_google(keyword: str, category: str):
    """구글 뉴스 RSS. API 키 불필요."""
    url = (
        "https://news.google.com/rss/search?q="
        + quote(keyword)
        + "&hl=ko&gl=KR&ceid=KR:ko"
    )
    out = []
    try:
        feed = feedparser.parse(url)
        for e in feed.entries:
            pub = None
            if getattr(e, "published_parsed", None):
                pub = dt.datetime(*e.published_parsed[:6], tzinfo=dt.timezone.utc)
            press = ""
            if getattr(e, "source", None) and e.source.get("title"):
                press = e.source.get("title")
            title = _clean(e.get("title"))
            # 구글은 제목 끝에 " - 언론사"를 붙임 → 분리
            if " - " in title and not press:
                title, press = title.rsplit(" - ", 1)
            out.append({
                "category": category,
                "keyword": keyword,
                "title": title,
                "press": press,
                "date": pub,
                "summary": _clean(e.get("summary")),
                "link": e.get("link", ""),
                "source": "구글뉴스",
            })
    except Exception as ex:
        print(f"  [구글 오류] {keyword}: {ex}")
    return out


def collect():
    """모든 키워드 수집 → 날짜필터 → 중복제거."""
    rows, seen = [], set()
    for category, words in KEYWORDS.items():
        for kw in words:
            print(f"· 검색 중: [{category}] {kw}")
            items = fetch_naver(kw, category, MAX_PER_KEYWORD)
            if not items:  # 네이버 키 없거나 결과 없으면 구글로 보완
                items = fetch_google(kw, category)
            for it in items:
                if not _within_range(it["date"], DAYS_BACK):
                    continue
                # 중복 키: 링크 우선, 없으면 제목
                key = (it["link"] or it["title"]).strip().lower()
                if not key or key in seen:
                    continue
                seen.add(key)
                rows.append(it)
            time.sleep(0.2)  # 과도한 요청 방지
    # 분류 → 날짜(최신순) 정렬
    rows.sort(key=lambda x: (x["category"], -(x["date"].timestamp() if x["date"] else 0)))
    return rows


# ---- 엑셀 저장 ----------------------------------------------------------

HEADERS = ["번호", "분류", "검색어", "제목", "언론사", "발행일시", "요약", "링크", "출처"]
WIDTHS = [6, 8, 14, 50, 16, 18, 60, 45, 10]


def _style_header(ws):
    fill = PatternFill("solid", fgColor="1F3864")          # 남색
    font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    side = Side(style="thin", color="BFBFBF")
    border = Border(left=side, right=side, top=side, bottom=side)
    for c, name in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill, cell.font, cell.border = fill, font, border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"


def save_excel(rows, path):
    sheet_name = dt.datetime.now(KST).strftime("%Y-%m-%d")
    if os.path.exists(path):
        wb = load_workbook(path)
        if sheet_name in wb.sheetnames:        # 같은 날 재실행 시 갱신
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name, 0)    # 최신 날짜를 맨 앞에
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    _style_header(ws)
    body_font = Font(name="맑은 고딕", size=9)
    side = Side(style="thin", color="D9D9D9")
    border = Border(left=side, right=side, top=side, bottom=side)
    wrap = Alignment(vertical="top", wrap_text=True)

    for i, it in enumerate(rows, 1):
        date_str = it["date"].astimezone(KST).strftime("%Y-%m-%d %H:%M") if it["date"] else "〔확인 필요〕"
        values = [i, it["category"], it["keyword"], it["title"], it["press"],
                  date_str, it["summary"], it["link"], it["source"]]
        r = i + 1
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font, cell.border, cell.alignment = body_font, border, wrap
            if c in (1, 2):
                cell.alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=r, column=8).hyperlink = it["link"] or None
        if it["link"]:
            ws.cell(row=r, column=8).font = Font(name="맑은 고딕", size=9, color="0563C1", underline="single")
    wb.save(path)
    return sheet_name, len(rows)


def main():
    load_external_config()
    print("=" * 50)
    print(" 카드·외환 뉴스 스크랩 시작")
    print(" 실행시각:", dt.datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S"))
    if not (NAVER_CLIENT_ID and NAVER_CLIENT_SECRET):
        print(" (네이버 키 미설정 → 구글 뉴스 RSS 사용)")
    print("=" * 50)

    rows = collect()
    if not rows:
        print("\n수집된 기사가 없습니다. DAYS_BACK 값을 늘리거나 키워드를 확인하세요.")
        return

    out_dir = OUTPUT_DIR if os.path.isabs(OUTPUT_DIR) else os.path.join(_base_dir(), OUTPUT_DIR)
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, OUTPUT_FILE)
    sheet, n = save_excel(rows, path)
    card = sum(1 for r in rows if r["category"] == "카드")
    fx = sum(1 for r in rows if r["category"] == "외환")
    print(f"\n완료 ▶ '{sheet}' 시트에 {n}건 저장 (카드 {card} / 외환 {fx})")
    print(f"파일: {os.path.abspath(path)}")


if __name__ == "__main__":
    main()
